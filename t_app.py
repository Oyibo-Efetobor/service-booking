import logging
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException, UnexpectedAlertPresentException
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ConversationHandler, ContextTypes

# Telegram conversation states
USERNAME, PASSWORD = range(2)

# Set up logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text('Welcome to Chapel Booking Bot!\n\nPlease enter your att username:')
    return USERNAME

async def get_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Delete the user's message for privacy
    try:
        await update.message.delete()
    except Exception:
        pass
    context.user_data['username'] = update.message.text
    await update.message.reply_text('Please enter your att password:')
    return PASSWORD

async def get_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    import openpyxl
    import os
    # Delete the user's message for privacy
    try:
        await update.message.delete()
    except Exception:
        pass
    context.user_data['password'] = update.message.text
    # Store chat_id and username in Excel (avoid duplicates)
    chat_id = update.effective_chat.id
    username = context.user_data.get('username', '')
    excel_path = os.path.join(os.path.dirname(__file__), 'bot_users.xlsx')
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        usernames = [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
        if username not in usernames:
            ws.append([chat_id, username])
            wb.save(excel_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['chat_id', 'username'])
        ws.append([chat_id, username])
        wb.save(excel_path)
    await update.message.reply_text('Attempting to log in and book your service...')
    # Start booking process
    result = await book_service(update, context)
    return ConversationHandler.END if result else USERNAME

async def book_service(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = context.user_data['username']
    password = context.user_data['password']
    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--start-maximized')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    booking_made = False  # Track if booking was made in this session
    try:
        driver.get('https://att2.lmu.edu.ng/log/login')
        time.sleep(2)
        driver.find_element(By.NAME, 'username').send_keys(username)
        driver.find_element(By.NAME, 'password').send_keys(password)
        driver.find_element(By.NAME, 'submit').click()
        time.sleep(3)
        # Check if login was successful
        if 'log/login' in driver.current_url:
            await update.message.reply_text('Username or password error. Please try again. \n\n Enter your username:')
            driver.quit()
            return False
        await update.message.reply_text('Login successful! Navigating to service schedule page...')
        # Step 3: Go to service schedule page with auto-reload until successful
        max_retries = 60
        for attempt in range(max_retries):
            driver.get('https://att2.lmu.edu.ng/check/serveChoice')
            time.sleep(5)
            if 'serveChoice' in driver.current_url:
                await update.message.reply_text(f'Arrived at service schedule page after {attempt+1} attempt(s).')
                break
            else:
                await update.message.reply_text(f'Attempt {attempt+1}: Not yet on service schedule page, reloading...')
        else:
            await update.message.reply_text('Failed to reach service schedule page after maximum retries.')
            driver.quit()
            return True
        # Step 4: Book 1st Service if available, else fallback to 2nd Service
        try:
            select_elem = driver.find_element(By.NAME, 'pudate')
            options = select_elem.find_elements(By.TAG_NAME, 'option')
            picked = False
            for option in options:
                if '1st Service' in option.text and 'Available Slots' in option.text:
                    option.click()
                    picked = True
                    await update.message.reply_text('Selected 1st Service. Trying to book...')
                    break
            if not picked:
                for option in options:
                    if '2nd Service' in option.text and 'Available Slots' in option.text:
                        option.click()
                        picked = True
                        await update.message.reply_text('Selected 2nd Service. Trying to book...')
                        break
            if not picked:
                await update.message.reply_text('No available service to pick.')
            else:
                # Step 2: Click "I agree" checkbox (by label text if needed)
                try:
                    agree_checkbox = driver.find_element(By.ID, 'confirm_remove_original')
                    if not agree_checkbox.is_selected():
                        driver.execute_script("arguments[0].click();", agree_checkbox)
                        time.sleep(1)
                except Exception as e:
                    await update.message.reply_text(f'Could not click the agreement checkbox: {e}')
                    driver.quit()
                    return True
                # Step 3: Handle the popup "Are you sure about your choice?" (alert) immediately after clicking checkbox
                try:
                    alert = driver.switch_to.alert
                    if 'Are you sure' in alert.text:
                        alert.accept()
                        await update.message.reply_text('Confirmed your choice in popup.')
                except Exception:
                    pass
                time.sleep(1)
                # Step 4: Click Reserve Seat button with robust retry logic
                from selenium.common.exceptions import ElementClickInterceptedException, UnexpectedAlertPresentException
                for i in range(5):
                    try:
                        # Accept any alert before clicking
                        try:
                            alert = driver.switch_to.alert
                            alert.accept()
                            time.sleep(1)
                        except Exception:
                            pass
                        try:
                            reserve_btn = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Reserve Seat']")
                        except NoSuchElementException:
                            try:
                                reserve_btn = driver.find_element(By.CSS_SELECTOR, "input.bg-emerald.fg-white[type='submit']")
                            except NoSuchElementException:
                                await update.message.reply_text(f'Attempt {i+1}: Reserve button not found, retrying...')
                                time.sleep(2)
                                continue
                        driver.execute_script("arguments[0].scrollIntoView(true);", reserve_btn)
                        time.sleep(0.5)
                        reserve_btn.click()
                        booking_made = True
                        await update.message.reply_text('Reservation attempted.')
                        # Accept any alert after clicking
                        try:
                            alert = driver.switch_to.alert
                            alert.accept()
                        except Exception:
                            pass
                        break
                    except UnexpectedAlertPresentException:
                        try:
                            alert = driver.switch_to.alert
                            alert.accept()
                        except Exception:
                            pass
                        time.sleep(1)
                    except ElementClickInterceptedException:
                        await update.message.reply_text(f'Attempt {i+1}: Reserve button not clickable yet, retrying...')
                        time.sleep(2)
                else:
                    await update.message.reply_text('Failed to click Reserve Seat after retries.')
        except NoSuchElementException as e:
            if 'pudate' in str(e):
                await update.message.reply_text('Service has already been booked previously for this account.')
                driver.quit()
                return True
            else:
                await update.message.reply_text(f'Could not complete booking: {e}')
        time.sleep(5)
        # Only take and send screenshot if booking was made in this session
        if booking_made:
            try:
                import base64
                screenshot = driver.execute_cdp_cmd("Page.captureScreenshot", {"fromSurface": True, "captureBeyondViewport": True, "format": "png"})
                await update.message.reply_photo(base64.b64decode(screenshot['data']), caption='Here is your booking screenshot.')
            except Exception as cdp_e:
                await update.message.reply_text(f'CDP full-page screenshot failed: {cdp_e}.')
        driver.quit()
        return True
    except Exception as e:
        # Check for DataTables warning or unexpected alert open error
        if (
            'DataTables warning' in str(e)
            or 'unexpected alert open' in str(e)
            or 'Cannot reinitialise DataTable' in str(e)
        ):
            await update.message.reply_text('Pick the necessary forms on your att portal to enable you book, then after click /start')
        else:
            await update.message.reply_text(f'Error: {e}')
        driver.quit()
        return True

def main():
    TOKEN = "8029449841:AAFXIqNoNgjM9Wn1T31NmHXLrjvebUFOh8A"
    app = ApplicationBuilder().token(TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_username)],
            PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_password)],
        },
        fallbacks=[]
    )
    app.add_handler(conv_handler)
    app.add_handler(CommandHandler('start', start))
    print('Bot is running. Press Ctrl+C to stop.')
    app.run_polling()

if __name__ == '__main__':
    main()
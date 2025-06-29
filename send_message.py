import os
import openpyxl
import asyncio
from telegram import Update, Bot
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes
from dotenv import load_dotenv

# Load tokens from .env or hardcode if needed
load_dotenv()
MAIN_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
BROADCAST_BOT_TOKEN = '7711565743:AAHjbLb5R-sgJUY2klQZxoWxEwlfkeDrYoM'

if not MAIN_BOT_TOKEN:
    MAIN_BOT_TOKEN = '8029449841:AAFXIqNoNgjM9Wn1T31NmHXLrjvebUFOh8A'  # fallback to your main bot token

# Path to user Excel file
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'bot_users.xlsx')

async def broadcast_to_all_users(update: Update):
    if not os.path.exists(EXCEL_PATH):
        print('No bot_users.xlsx file found.')
        return
    
    print("Starting broadcast...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    bot = Bot(token=MAIN_BOT_TOKEN)  # Use main bot token for sending messages
    
    total_users = sum(1 for row in ws.iter_rows(min_row=2))
    success_count = 0
    fail_count = 0
    
    text = update.message.text
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        chat_id = row[0]
        try:
            await bot.send_message(chat_id=chat_id, text=text)
            success_count += 1
            print(f"Successfully sent message to user {chat_id}")
        except Exception as e:
            fail_count += 1
            print(f"Failed to send message to {chat_id}: {str(e)}")
        
        # Add a small delay between messages to avoid rate limiting
        await asyncio.sleep(0.1)
    
    print(f"\nBroadcast completed:")
    print(f"Total users: {total_users}")
    print(f"Successful: {success_count}")
    print(f"Failed: {fail_count}")
    return success_count, fail_count

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Only allow the admin (the creator of the broadcast bot) to send broadcasts
    admin_id = update.effective_user.id
    # Optionally, restrict to your Telegram user ID for security
    # if admin_id != YOUR_TELEGRAM_USER_ID:
    #     await update.message.reply_text('You are not authorized to broadcast.')
    #     return
    
    await broadcast_to_all_users(update)
    await update.message.reply_text('Broadcasting message to all users...')

async def main():
    # This is the broadcast bot's token
    BROADCAST_BOT_TOKEN = '7711565743:AAHjbLb5R-sgJUY2klQZxoWxEwlfkeDrYoM'
    app = ApplicationBuilder().token(BROADCAST_BOT_TOKEN).build()
      # Handle text messages only
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    print('Broadcast bot is running. Send a message, photo, video, or document to broadcast to all users.')
    await app.run_polling()

if __name__ == '__main__':
    import sys
    import asyncio
    import nest_asyncio

    if sys.platform.startswith('win') and sys.version_info >= (3, 8):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    # Apply nest_asyncio to allow nested event loops
    nest_asyncio.apply()
    
    # Create and run event loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    finally:
        loop.close()
